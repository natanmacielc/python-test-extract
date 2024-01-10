import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook


class ExcelCreator:
    _filename: str = '/app/test-data/test-data.xlsx'

    @staticmethod
    def create_excel(data: pd.DataFrame):
        workbook = Workbook()
        worksheet = workbook.active

        for idx, col in enumerate(data.columns):
            cell = worksheet.cell(row=1, column=idx + 1)
            cell.value = col  # Adicionando os nomes das colunas
            cell.fill = PatternFill(start_color='0000ff', end_color='0000ff', fill_type='solid')

        for r, row in enumerate(data.values, start=2):  # Iniciando da linha 2 para os dados
            for c, val in enumerate(row, start=1):
                worksheet.cell(row=r, column=c).value = val

        workbook.save(ExcelCreator._filename)
        print(f'Os dados foram exportados para {ExcelCreator._filename} com sucesso!')
